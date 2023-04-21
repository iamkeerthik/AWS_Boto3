import boto3
import openpyxl
import os

# Set the AWS profile and region to use
session = boto3.Session(profile_name='Stage', region_name='us-east-2')

# Create an EC2 client using the specified profile and region
ec2 = session.client('ec2')

# Get a list of all volumes in the account
response = ec2.describe_volumes()

# Create an Excel spreadsheet if it doesn't already exist
if not os.path.exists('volumes.xlsx'):
    workbook = openpyxl.Workbook()
    workbook.save('volumes.xlsx')

# Open the Excel spreadsheet
workbook = openpyxl.load_workbook('volumes.xlsx')
worksheet = workbook.active

# Write the headers to the spreadsheet if it's empty
if worksheet.max_row == 1 and worksheet.max_column == 1:
    worksheet.cell(row=1, column=1).value = 'Volume ID'
    worksheet.cell(row=1, column=2).value = 'Volume Name'
    worksheet.cell(row=1, column=3).value = 'Volume Type'
    worksheet.cell(row=1, column=4).value = 'Volume Size (GB)'
    worksheet.cell(row=1, column=5).value = 'Creation Date'
    worksheet.cell(row=1, column=6).value = 'Attached Instance ID'
    worksheet.cell(row=1, column=7).value = 'Availability'
    worksheet.cell(row=1, column=8).value = 'Snapshot ID'

# Loop through each volume and update the spreadsheet
for index, volume in enumerate(response['Volumes']):
    volume_id = volume['VolumeId']
    volume_name = ''
    if 'Tags' in volume:
        for tag in volume['Tags']:
            if tag['Key'] == 'Name':
                volume_name = tag['Value']
                break
    volume_type = volume['VolumeType']
    volume_size = volume['Size']
    creation_date = volume['CreateTime'].strftime('%Y-%m-%d %H:%M:%S')
    instance_id = ''
    if 'Attachments' in volume and volume['Attachments']:
        instance_id = volume['Attachments'][0]['InstanceId']
    available = 'Available' if volume['State'] == 'available' else 'Not available'
    snapshot_id = ''
    if 'SnapshotId' in volume:
        snapshot_id = volume['SnapshotId']
    worksheet.cell(row=index+2, column=1).value = volume_id
    worksheet.cell(row=index+2, column=2).value = volume_name
    worksheet.cell(row=index+2, column=3).value = volume_type
    worksheet.cell(row=index+2, column=4).value = volume_size
    worksheet.cell(row=index+2, column=5).value = creation_date
    worksheet.cell(row=index+2, column=6).value = instance_id
    worksheet.cell(row=index+2, column=7).value = available
    worksheet.cell(row=index+2, column=8).value = snapshot_id

# Save the changes to the spreadsheet
workbook.save('volumes.xlsx')
