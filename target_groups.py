import boto3
import openpyxl
from botocore.exceptions import ClientError
# Enter the names of the regions to check
region_names = ['us-west-2']

# Enter the name of the AWS profile to use
profile_name = 'Stage'

# Enter the name of the Excel file to create or update
excel_file = 'target_groups-oregon.xlsx'

# Enter the name of the worksheet to create or update
worksheet_name = 'Target Groups'

# Create a list to store the number of targets in each target group
target_counts = []

# Create an EC2 client object using the specified profile
session = boto3.Session(profile_name=profile_name)
elbv2 = session.client('elbv2')

# Loop through each region
for region_name in region_names:
    print(f"Scanning region {region_name}...")
    # Create an ELBV2 client object for the current region
    elbv2_region = session.client('elbv2', region_name=region_name)
    # Get a list of all target groups in the current region
    target_groups = elbv2_region.describe_target_groups()['TargetGroups']
    # Loop through each target group
    for target_group in target_groups:
        # Get the target group ARN
        target_group_arn = target_group['TargetGroupArn']
        # Get a list of all targets in the target group
        targets = elbv2_region.describe_target_health(TargetGroupArn=target_group_arn)['TargetHealthDescriptions']
        # Add the number of targets in the target group to the target_counts list
        target_counts.append(len(targets))

# Create a new Excel workbook object
workbook = openpyxl.Workbook()

# Create a worksheet for the target groups
worksheet = workbook.create_sheet(title=worksheet_name)

# Add column headers to the worksheet
worksheet['A1'] = 'Target Group Name'
worksheet['B1'] = 'Number of Targets'
worksheet['C1'] = 'Associated with Load Balancer'
worksheet['D1'] = 'Loadbalaner Name'

# Loop through each region again to add target group information to the worksheet
current_row = 2
for region_name in region_names:
    # Create an ELBV2 client object for the current region
    elbv2_region = session.client('elbv2', region_name=region_name)
    # Get a list of all target groups in the current region
    target_groups = elbv2_region.describe_target_groups()['TargetGroups']
    # Loop through each target group
    for target_group in target_groups:
        # Get the target group name
        target_group_name = target_group['TargetGroupName']
        # Get the number of targets in the target group
        target_count = target_counts.pop(0)
        # Check if the target group is associated with a load balancer
        lb_associated = 'Yes' if target_group['LoadBalancerArns'] else 'No'
        # Get the load balancer names associated with the target group, if any
        lb_names = []
        if target_group['LoadBalancerArns']:
            lb_arns = target_group['LoadBalancerArns']
            for lb_arn in lb_arns:
                try:
                    lb_name = elbv2_region.describe_load_balancers(LoadBalancerArns=[lb_arn])['LoadBalancers'][0]['LoadBalancerName']
                    lb_names.append(lb_name)
                except ClientError as e:
                    if e.response['Error']['Code'] == 'ValidationError':
                        print(f"Error: {lb_arn} is not a valid load balancer ARN")
         # Concatenate the list of load balancer names into a single string
        lb_names_str = ', '.join(lb_names)
        # Add the target group name, number of targets, and associated with load balancer status to the worksheet
        worksheet.cell(row=current_row, column=1, value=target_group_name)
        worksheet.cell(row=current_row, column=2, value=target_count)
        worksheet.cell(row=current_row, column=3, value=lb_associated)
        worksheet.cell(row=current_row, column=4, value=lb_names_str)
        # Increment the current row counter
        current_row += 1

# Save the workbook to the specified Excel file
workbook.save(excel_file)

print("Target group information has been added to the Excel file.")
